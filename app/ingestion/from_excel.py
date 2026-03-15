from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from app.utils.hashing import sha256_text
from app.utils.text_cleaning import flatten_text, normalize_punctuation, normalize_whitespace


INPUT_XLSX = Path('SOAS_profiles.xlsx')
OUTPUT_DIR = Path('data/processed_profiles')
SHEET_NAME = 'AcademicProfiles'


LANGUAGE_CANDIDATES = [
    'English',
    'Arabic',
    'French',
    'Mandarin',
    'Chinese',
    'Japanese',
    'Hindi',
    'Urdu',
    'Bengali',
    'Spanish',
    'German',
    'Italian',
    'Persian',
    'Turkish',
    'Russian',
    'Korean',
    'Portuguese',
    'Punjabi',
    'Tamil',
    'Greek',
]


def slugify(text: str) -> str:
    text = str(text).strip().lower()
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')


def clean_title(text: str | None) -> str | None:
    cleaned = normalize_whitespace(text)
    if not cleaned:
        return None

    lines = [line.strip(' -*;:') for line in cleaned.split('\n')]
    lines = [line for line in lines if line]

    unique_lines: list[str] = []
    seen: set[str] = set()
    for line in lines:
        key = line.lower()
        if key not in seen:
            unique_lines.append(line)
            seen.add(key)

    if not unique_lines:
        return None
    return ' | '.join(unique_lines)


def clean_department(text: str | None) -> str | None:
    return flatten_text(text)


def split_concatenated_keywords(text: str) -> str:
    text = text.strip()
    text = re.sub(r'(?<=[a-z])(?=[A-Z])', '; ', text)
    text = re.sub(r'(?<=[A-Z])(?=[A-Z][a-z])', '; ', text)
    text = text.replace('|', ';')
    text = text.replace('/', '; ')
    return text


def split_keywords(value: str | None) -> list[str]:
    if not value:
        return []

    text = normalize_whitespace(value)
    if not text:
        return []

    text = split_concatenated_keywords(text)
    parts = re.split(r';|\n|,|\u2022', text)

    cleaned: list[str] = []
    seen: set[str] = set()
    for part in parts:
        normalized = normalize_punctuation(part)
        normalized = re.sub(r'\s+', ' ', normalized).strip(' .,-;:')
        if not normalized or len(normalized) < 2:
            continue
        key = normalized.lower()
        if key not in seen:
            cleaned.append(normalized)
            seen.add(key)

    return cleaned


def extract_languages(text: str | None) -> list[str]:
    if not text:
        return []

    text_lower = text.lower()
    found: list[str] = []
    for lang in LANGUAGE_CANDIDATES:
        if re.search(rf'\b{re.escape(lang.lower())}\b', text_lower):
            found.append(lang)
    return found


def clean_links(value: str | None) -> str | None:
    text = normalize_whitespace(value)
    if not text:
        return None
    return re.sub(r'\n+', '\n', text)


def clean_biography(text: str | None) -> str | None:
    flattened = flatten_text(text)
    if not flattened:
        return None
    return normalize_punctuation(flattened) or None


def clean_research_interests(text: str | None) -> str | None:
    normalized = normalize_whitespace(text)
    if not normalized:
        return None

    normalized = split_concatenated_keywords(normalized)
    if ';' in normalized or '\n' in normalized:
        parts = split_keywords(normalized)
        if parts:
            return '; '.join(parts)

    flattened = flatten_text(normalized)
    if not flattened:
        return None
    return normalize_punctuation(flattened)


def build_profile(row: dict) -> dict:
    name = flatten_text(row.get('Full name'))
    title = clean_title(row.get('Job title'))
    department = clean_department(row.get('Department'))
    bio = clean_biography(row.get('Bio (from SOAS profile)'))
    email_contact = flatten_text(row.get('Email/contact (as published)'))
    research_keywords = clean_research_interests(row.get('Research interests/keywords'))
    research_links = clean_links(row.get('Research links (centres/projects/publications/media)'))
    profile_url = flatten_text(row.get('Profile URL'))
    snapshot_date = flatten_text(row.get('Snapshot date'))
    last_checked = flatten_text(row.get('Last checked')) or snapshot_date or '2026-03-08'
    content_hash = flatten_text(row.get('Content hash (optional)'))
    notes = normalize_whitespace(row.get('Notes / data quality flags'))

    if not name or not profile_url:
        raise ValueError('Missing required name or profile_url')

    profile_id = f'soas-{slugify(name)}'
    expertise_topics = split_keywords(research_keywords)

    languages = extract_languages(
        ' '.join(
            filter(
                None,
                [bio, research_keywords, research_links, title, department],
            )
        )
    )

    hash_source = ' | '.join(
        filter(
            None,
            [
                name,
                title or '',
                department or '',
                bio or '',
                research_keywords or '',
                research_links or '',
                profile_url,
                last_checked,
            ],
        )
    )
    final_hash = content_hash or sha256_text(hash_source)

    profile = {
        'profile_id': profile_id,
        'name': name,
        'title': title,
        'department': department,
        'expertise_topics': expertise_topics,
        'biography': bio,
        'research_interests': research_keywords,
        'publications': research_links,
        'languages': languages,
        'source_url': profile_url,
        'last_checked': last_checked,
        'content_hash': final_hash,
    }

    if email_contact:
        profile['email_contact'] = email_contact
    if notes:
        profile['notes'] = notes

    return profile


def main() -> None:
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f'Input workbook not found: {INPUT_XLSX}')

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Found: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    written = 0
    skipped = 0

    for raw_row in rows[1:]:
        row = dict(zip(headers, raw_row))
        if not any(v is not None and str(v).strip() for v in raw_row):
            continue

        try:
            profile = build_profile(row)
            output_path = OUTPUT_DIR / f"{profile['profile_id']}.json"
            output_path.write_text(
                json.dumps(profile, ensure_ascii=False, indent=2),
                encoding='utf-8',
            )
            written += 1
        except Exception as exc:
            skipped += 1
            print(f"Skipped row for '{row.get('Full name')}' -> {exc}")

    print(f'Done. Wrote {written} JSON profiles to {OUTPUT_DIR}')
    if skipped:
        print(f'Skipped {skipped} rows')


if __name__ == '__main__':
    main()
